"""
FastAPI application for the ICAI NCE Conversion Tool.

Routes:
  GET  /                 -> single-page UI
  POST /api/extract      -> upload files + constitution -> structured JSON (Claude)
  POST /api/generate     -> structured JSON -> formula-linked .xlsx download
  GET  /health           -> health check

The /api/generate endpoint works WITHOUT an API key (pure engine), so the manual
or review-then-generate path needs no Claude. /api/extract requires ANTHROPIC_API_KEY.
"""
from __future__ import annotations
import io, json, os, urllib.request, urllib.parse
from dataclasses import asdict
from typing import List

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Header
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles

from models import Payload
from builder import build_workbook

app = FastAPI(title="ICAI NCE Conversion Tool", version="1.0.0")


# In ACCESS_CODES each entry is "code" or "code:limit" (limit 0/absent = unlimited),
# e.g. "MSjoshi@725:0,Demo@12345:3". Codes must not contain a ':'.
def parse_codes():
    raw = os.environ.get("ACCESS_CODES", "genius2025")
    out = {}
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        if ":" in part:
            code, _, lim = part.rpartition(":")
            code = code.strip()
            try:
                limit = max(0, int(lim.strip()))   # explicit per-code limit (0 = unlimited)
            except ValueError:
                code, limit = part, None            # ':' was part of the code, not a limit
        else:
            code, limit = part, None                # no explicit limit -> use DEFAULT
        if code:
            out[code] = limit
    return out


def default_limit():
    try:
        return max(0, int(os.environ.get("DEFAULT_CONVERSION_LIMIT", "0") or 0))
    except ValueError:
        return 0


def valid_codes():
    return set(parse_codes().keys())


def require_code(code):
    if not code or code.strip() not in valid_codes():
        raise HTTPException(401, "Invalid or inactive access code. Please contact "
                                 "M S Joshi & Co. (connect@msjc.in) for access.")
    return code.strip()


# ---- usage counter (persistent via Upstash Redis if configured; else in-memory) ----
_MEM_USAGE = {}


def _upstash():
    url = os.environ.get("UPSTASH_REDIS_REST_URL")
    tok = os.environ.get("UPSTASH_REDIS_REST_TOKEN")
    return (url.rstrip("/"), tok) if url and tok else None


def _redis(path):
    cfg = _upstash()
    if not cfg:
        return None
    url, tok = cfg
    req = urllib.request.Request(f"{url}/{path}", headers={"Authorization": f"Bearer {tok}"})
    with urllib.request.urlopen(req, timeout=6) as r:
        return json.loads(r.read().decode()).get("result")


def _key(code):
    return "nceusage:" + urllib.parse.quote(code, safe="")


def usage_get(code):
    if _upstash():
        try:
            v = _redis("get/" + _key(code))
            return int(v) if v not in (None, "") else 0
        except Exception:
            pass
    return _MEM_USAGE.get(code, 0)


def usage_incr(code):
    if _upstash():
        try:
            return int(_redis("incr/" + _key(code)))
        except Exception:
            pass
    _MEM_USAGE[code] = _MEM_USAGE.get(code, 0) + 1
    return _MEM_USAGE[code]


def code_status(code):
    explicit = parse_codes().get(code)
    limit = explicit if explicit is not None else default_limit()
    used = usage_get(code)
    return {"limit": limit, "used": used,
            "remaining": (None if limit <= 0 else max(0, limit - used)),
            "unlimited": limit <= 0}

HERE = os.path.dirname(__file__)
STATIC_DIR = HERE
if os.path.isdir(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


@app.api_route("/health", methods=["GET", "HEAD"])
def health():
    store = "memory"
    if _upstash():
        try:
            _redis("get/nce_health_probe")   # harmless read to confirm connectivity
            store = "redis"
        except Exception:
            store = "memory (redis unreachable)"
    return {"status": "ok",
            "llm_configured": bool(os.environ.get("ANTHROPIC_API_KEY")),
            "codes_configured": bool(os.environ.get("ACCESS_CODES")),
            "usage_store": store}


@app.get("/", response_class=HTMLResponse)
def index():
    path = os.path.join(STATIC_DIR, "index.html")
    if os.path.exists(path):
        return HTMLResponse(open(path, encoding="utf-8").read())
    return HTMLResponse("<h1>ICAI NCE Conversion Tool</h1><p>UI not found.</p>")


@app.post("/api/verify")
def api_verify(x_access_code: str = Header(None)):
    code = require_code(x_access_code)
    return {"valid": True, **code_status(code)}


@app.post("/api/admin/reset")
async def api_admin_reset(target: str = Form(...), x_admin_key: str = Header(None)):
    admin = os.environ.get("ADMIN_KEY")
    if not admin or x_admin_key != admin:
        raise HTTPException(401, "Admin key required.")
    if _upstash():
        try:
            _redis("del/" + _key(target))
        except Exception:
            pass
    _MEM_USAGE.pop(target, None)
    return {"reset": target, "used": usage_get(target)}


@app.post("/api/extract")
def api_extract(constitution: str = Form(...),
                      extra: str = Form(""),
                      files: List[UploadFile] = File(...),
                      x_access_code: str = Header(None)):
    code = require_code(x_access_code)
    if constitution not in ("proprietorship", "partnership"):
        raise HTTPException(400, "constitution must be 'proprietorship' or 'partnership'")
    if not os.environ.get("ANTHROPIC_API_KEY"):
        raise HTTPException(503, "ANTHROPIC_API_KEY is not configured on the server. "
                                 "Use the manual JSON path, or set the key and redeploy.")
    payload_files = []
    for f in files:
        nm = (f.filename or "").lower()
        if nm.endswith(".doc") and not nm.endswith(".docx"):
            raise HTTPException(415, "This looks like a legacy .doc file (old Word format). "
                                     "Please open it in Word and use File > Save As > Word Document (.docx), "
                                     "then upload the .docx. PDF and Excel are also supported.")
        payload_files.append((f.filename, f.file.read()))
    # enforce per-code conversion limit (extraction is the token-cost step)
    st = code_status(code)
    if not st["unlimited"] and st["used"] >= st["limit"]:
        raise HTTPException(429, f"This access code has reached its limit of {st['limit']} "
                                 "conversion(s). Please contact M S Joshi & Co. "
                                 "(connect@msjc.in) to renew or upgrade.")
    usage_incr(code)
    try:
        from llm import extract_payload, reconcile_and_correct
        import reconcile as rec
        payload = extract_payload(constitution, payload_files, extra)
        payload, discrepancies, fixes = reconcile_and_correct(constitution, payload_files, payload)
    except Exception as e:
        raise HTTPException(500, f"Extraction failed: {e}")
    return JSONResponse({
        "payload": asdict(payload),
        "usage": code_status(code),
        "reconciliation": {
            "passed": len(discrepancies) == 0,
            "discrepancies": discrepancies,
            "fixes": fixes,
            "report": rec.report_text(discrepancies),
        },
    })


def _add_review_sheet(data: bytes, discrepancies) -> bytes:
    """Prepend a red 'Review' worksheet flagging every line that does not yet tie to
    source, so the draft workbook is usable and the CA sees exactly what to check."""
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.create_sheet("\u26A0 Review", 0)
    ws.sheet_properties.tabColor = "C0392B"
    F = "Calibri Light"
    ws["A1"] = ("Auto-generated draft - items to verify. These figures do not yet tie exactly to "
                "the source; please review the flagged lines below and correct them before finalising or signing.")
    ws["A1"].font = Font(name=F, size=12, bold=True, color="C0392B")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A1:E1"); ws.row_dimensions[1].height = 34
    hdr = ["Check", "Year", "Expected (source)", "Got (converted)", "Difference"]
    ws.append([]); ws.append(hdr)
    fill = PatternFill("solid", fgColor="F2DEDE")
    thin = Side(style="thin", color="D9D9D9"); bord = Border(left=thin,right=thin,top=thin,bottom=thin)
    for c in range(1, 6):
        cell = ws.cell(row=3, column=c)
        cell.font = Font(name=F, size=11, bold=True); cell.fill = fill; cell.border = bord
    r = 4
    for x in discrepancies:
        ws.cell(row=r, column=1, value=x.get("check", ""))
        ws.cell(row=r, column=2, value=x.get("year", ""))
        ws.cell(row=r, column=3, value=round(float(x.get("expected", 0) or 0), 2))
        ws.cell(row=r, column=4, value=round(float(x.get("got", 0) or 0), 2))
        ws.cell(row=r, column=5, value=round(float(x.get("diff", 0) or 0), 2))
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c); cell.font = Font(name=F, size=11); cell.border = bord
            if c >= 3: cell.number_format = '#,##0.00;(#,##0.00)'
        r += 1
    for col, w in zip("ABCDE", [52, 16, 20, 20, 18]):
        ws.column_dimensions[col].width = w
    out = io.BytesIO(); wb.save(out); return out.getvalue()


@app.post("/api/generate")
def api_generate(payload: dict, denomination: str = "actual", x_access_code: str = Header(None)):
    require_code(x_access_code)
    import reconcile as rec
    try:
        p = Payload.parse(payload)
    except Exception as e:
        raise HTTPException(400, f"Invalid payload: {e}")
    discrepancies = rec.reconcile(p)
    try:
        data = build_workbook(p, denomination)
    except Exception as e:
        raise HTTPException(400, f"Build failed: {e}")
    # DRAFT-NOW: always produce the workbook. If anything does not tie, prepend a
    # red "Review" tab listing the exact lines to verify, so the reviewing CA can
    # correct the one figure in seconds instead of being blocked entirely.
    if discrepancies:
        data = _add_review_sheet(data, discrepancies)
    name = (p.entity.name or "Entity").replace(" ", "_")
    fy = p.entity.cy_fy or "FY"
    fname = f"{name}_FS_FY{fy}_ICAI_NCE.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
