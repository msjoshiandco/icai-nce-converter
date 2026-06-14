"""
openpyxl workbook engine for the ICAI NCE format.

Consumes a models.Payload and produces a formula-linked .xlsx:
  Index · Balance Sheet · Statement of P&L (+Appropriation for a firm) · Notes

Guarantees implemented here:
  - Dynamic note suppression (Nil in both years) + contiguous Schedule III numbering
  - Every BS/P&L face figure is a cross-sheet formula to its note total
  - Leaf note lines are the only hardcoded numbers
  - Capital closing(s) computed by formula and linked to the BS
  - PPE transposed schedule (Gross / Acc-Dep / Net), no deduction columns
  - Calibri Light 11 on every cell incl. empty (styles.xml + theme patch)
  - No underline anywhere
"""
from __future__ import annotations
import io, zipfile, re
from typing import Dict, Tuple, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.page import PageMargins

from models import Payload, Note, OwnerCapital, SUB_KINDS

FONT_NAME = "Calibri Light"
# Indian digit grouping (lakh/crore). LibreOffice/Excel only honour grouping via
# explicit [>=...] conditions, and allow at most two conditions per format, so we
# pick the format per cell by sign: positives/formulas -> POS (crore-aware),
# negatives -> NEG (bracketed), exact zero -> dash.
POS_FMT = r'[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00'
NEG_FMT = r'[<=-10000000](##\,##\,##\,##0.00);[<=-100000](##\,##\,##0.00);(##,##0.00)'
ZERO_FMT = '"-"'
NUMFMT = POS_FMT

# ---- Schedule III mapping ---------------------------------------------------
# Balance-sheet line plan: (section_header, [(sub_label, note_key), ...])
BS_LIABILITIES = [
    ("1  Owners' Funds", [
        ("(a)  Owners'/Partners' Capital Account", "capital"),
        ("(b)  Reserves and Surplus", "reserves"),
    ]),
    ("2  Non-current liabilities", [
        ("(a)  Long-term borrowings", "lt_borrowings"),
        ("(b)  Deferred tax liabilities (Net)", "dtl"),
        ("(c)  Other long-term liabilities", "other_lt_liab"),
        ("(d)  Long-term provisions", "lt_provisions"),
    ]),
    ("3  Current liabilities", [
        ("(a)  Short-term borrowings", "st_borrowings"),
        ("(b)  Trade payables", "trade_payables"),
        ("(c)  Other current liabilities", "other_cl"),
        ("(d)  Short-term provisions", "st_provisions"),
    ]),
]
BS_ASSETS = [
    ("1  Non-current assets", [
        ("(a)  Property, Plant and Equipment", "ppe"),
        ("(b)  Non-current investments", "nc_investments"),
        ("(c)  Deferred tax assets (Net)", "dta"),
        ("(d)  Long-term loans and advances", "lt_loans_adv"),
        ("(e)  Other non-current assets", "other_nca"),
    ]),
    ("2  Current assets", [
        ("(a)  Current investments", "current_investments"),
        ("(b)  Inventories", "inventories"),
        ("(c)  Trade receivables", "trade_receivables"),
        ("(d)  Cash and bank balances", "cash_bank"),
        ("(e)  Short-term loans and advances", "st_loans_adv"),
        ("(f)  Other current assets", "other_ca"),
    ]),
]

PL_EXPENSE_PLAN = [
    ("(a)  Cost of materials consumed / Purchases of stock-in-trade", "cost_materials"),
    ("(b)  Changes in inventories", "changes_inventory"),
    ("(c)  Employee benefits expense", "employee_benefits"),
    ("(d)  Finance costs", "finance_costs"),
    ("(e)  Depreciation and amortisation expense", "depreciation"),
    ("(f)  Other expenses", "other_expenses"),
]

# Notes whose presence does not depend on a numeric balance
ALWAYS_RETAINED = {"entity", "policies", "capital", "prev_year", "rounding"}
# Schedule III canonical order for numbering
SCHEDULE_ORDER = [
    "entity", "policies", "capital",
    "reserves", "lt_borrowings", "dtl", "other_lt_liab", "lt_provisions",
    "st_borrowings", "trade_payables", "other_cl", "st_provisions",
    "ppe", "nc_investments", "dta", "lt_loans_adv", "other_nca",
    "current_investments", "inventories", "trade_receivables",
    "cash_bank", "st_loans_adv", "other_ca",
    "revenue", "other_income", "cost_materials", "changes_inventory",
    "employee_benefits", "finance_costs", "depreciation", "other_expenses",
    "contingent", "related_party", "segment", "msmed", "forex",
    "confirmation", "prev_year", "rounding",
]

THIN = Side(style="thin")
DOUBLE = Side(style="double")

# presentation fills (modelled on the firm's company-format statements)
HEADER_FILL = PatternFill("solid", fgColor="DED8C6")   # warm tan column/table-header band
CY_FILL = PatternFill("solid", fgColor="EBF1EA")        # faint current-year column highlight


class Engine:
    def __init__(self, payload: Payload, denomination: str = "actual"):
        self.p = payload
        self.wb = Workbook()
        self.firm = payload.entity.constitution == "partnership"
        self.anchor: Dict[str, str] = {}
        self.note_no: Dict[str, int] = {}
        self.retained: List[str] = []
        self.numcell: Dict[str, str] = {}     # note key -> "AD{row}" on Notes
        # (value_scale, display_suffix, caption_word, threshold_multiplier)
        cfg = {
            "actual":    (1.0,  "",   "",             1),
            "thousands": (1.0,  ",",  "in thousands",  1000),
            "millions":  (1.0,  ",,", "in millions",   1000000),
            "lakhs":     (1e-5, "",   "in lakhs",      1),
            "crores":    (1e-7, "",   "in crores",     1),
        }.get((denomination or "actual").lower(), (1.0, "", "", 1))
        self.denom_scale, self.denom_suffix, self.denom_word, _D = cfg
        # thousands/millions scale DISPLAY only (cells stay absolute); lakhs/crores scale the value.
        # Thresholds are multiplied by _D so the Indian grouping matches the DISPLAYED magnitude.
        suf = self.denom_suffix
        cr, lk = int(1e7 * _D), int(1e5 * _D)
        self.POS = (f'[>={cr}]##\\,##\\,##\\,##0.00{suf};'
                    f'[>={lk}]##\\,##\\,##0.00{suf};##,##0.00{suf}')
        self.NEG = (f'[<=-{cr}](##\\,##\\,##\\,##0.00{suf});'
                    f'[<=-{lk}](##\\,##\\,##0.00{suf});(##,##0.00{suf})')
        self.ZERO = '"-"' 

    # -- styling helpers ------------------------------------------------------
    def _f(self, bold=False, italic=False, size=11, color=None):
        return Font(name=FONT_NAME, size=size, bold=bold, italic=italic,
                    underline=None, color=color)

    def cell(self, ws, r, c, val=None, bold=False, italic=False, num=False,
             center=False, right=False, wrap=False, size=11, top=False,
             bottom=False, double_bottom=False, fill=None, val_hint=None):
        if isinstance(val, str) and "\u20b9" in val:
            val = val.replace("\u20b9", "Rs. ")
        if num and isinstance(val, (int, float)):
            val = val * self.denom_scale
        cl = ws.cell(row=r, column=c, value=val)
        cl.font = self._f(bold, italic, size)
        al = Alignment(vertical="top", wrap_text=wrap,
                       horizontal="center" if center else ("right" if right else None))
        cl.alignment = al
        if fill is not None:
            cl.fill = fill
        if num:
            hint = val if isinstance(val, (int, float)) else (
                val_hint * self.denom_scale if isinstance(val_hint, (int, float)) else None)
            if isinstance(hint, (int, float)):
                cl.number_format = self.ZERO if abs(hint) < 0.005 else (self.NEG if hint < 0 else self.POS)
            else:
                cl.number_format = self.POS
        b = {}
        if top: b["top"] = THIN
        if bottom: b["bottom"] = THIN
        if double_bottom: b["bottom"] = DOUBLE
        if b:
            cl.border = Border(**b)
        return cl

    # -- suppression & numbering ---------------------------------------------
    def compute_retained(self):
        has_ppe = bool(self.p.ppe_assets)
        present = {n.key for n in self.p.notes if not n.is_nil()}
        for key in SCHEDULE_ORDER:
            keep = False
            if key in ALWAYS_RETAINED:
                keep = True
            elif key in ("ppe", "depreciation") and has_ppe:
                keep = True
            elif key == "st_provisions" and self.firm:
                keep = True  # firm tax provision
            elif key in present:
                keep = True
            if keep:
                self.retained.append(key)
        for i, key in enumerate(self.retained, start=1):
            self.note_no[key] = i

    # ========================================================================
    def build(self) -> bytes:
        self.compute_retained()
        self.wb.remove(self.wb.active)
        self.ws_notes = self.wb.create_sheet("Notes")
        self.ws_bs = self.wb.create_sheet("Balance Sheet")
        self.ws_pl = self.wb.create_sheet("Statement of P&L")
        self.ws_idx = self.wb.create_sheet("Index")
        self.ws_cap = None   # capital account is now rendered inline within Notes
        self.ws_ppe = self.wb.create_sheet("PPE Schedule") if "ppe" in self.note_no else None
        self._build_numbering()
        self._denomination_footnote()
        if self.ws_ppe is not None:
            self.build_ppe_sheet(self.note_no["ppe"])
        self.build_notes()
        self.build_balance_sheet()
        self.build_pl()
        self.build_index()
        self._highlight_cy(self.ws_bs, 4, 4)
        self._highlight_cy(self.ws_pl, 4, 4)
        self._highlight_cy(self.ws_notes, 4, 5)
        for ws in (self.ws_idx, self.ws_bs, self.ws_pl, self.ws_notes):
            self._a4(ws)
            ws.print_area = f"A1:F{ws.max_row}"
        if self.ws_ppe is not None:
            self.ws_ppe.page_setup.paperSize = 9
            self.ws_ppe.print_area = f"A1:J{self.ws_ppe.max_row}"
        self._reorder()
        self._no_gridlines()
        return self._save_with_theme_patch()

    def _reorder(self):
        order = ["Index", "Balance Sheet", "Statement of P&L", "Notes", "PPE Schedule"]
        self.wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

    def _highlight_cy(self, ws, col, r1):
        """Highlight only the current-year FIGURE cells (numbers / formulas),
        leaving heading, table-header, footnote and sub-note text rows un-shaded."""
        for rr in range(r1, ws.max_row + 1):
            v = ws.cell(row=rr, column=col).value
            if isinstance(v, (int, float)) or (isinstance(v, str) and v.startswith("=")):
                ws.cell(row=rr, column=col).fill = CY_FILL

    def _a4(self, ws):
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = 9
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2)

    def _scaled(self, fmt):
        return fmt.replace("0.00", "0.00" + self.denom_suffix) if self.denom_suffix else fmt

    def _amount_caption(self):
        if self.denom_word:
            return f"(Amount in Rs. {self.denom_word}, except as otherwise stated)"
        return "(Amount in Rs., except as otherwise stated)"

    def _build_numbering(self):
        col = 30  # column AD on the Notes sheet (hidden) holds the live note numbers
        for i, key in enumerate(self.retained):
            row = i + 1
            cl = self.ws_notes.cell(row=row, column=col,
                                    value=(1 if i == 0 else f"=AD{row-1}+1"))
            cl.font = self._f()
            self.numcell[key] = f"AD{row}"
        self.ws_notes.column_dimensions["AD"].hidden = True

    def _heading_formula(self, key):
        cell = self.numcell.get(key)
        return f'="Note "&{cell}' if cell else f"Note {self.note_no.get(key, '')}"

    def _subnum(self, key, k):
        cell = self.numcell.get(key)
        return f'={cell}&".{k}"' if cell else f"{self.note_no.get(key, '')}.{k}"

    def _denomination_footnote(self):
        if not self.denom_word:
            return
        rn = self.p.note("rounding")
        if rn is None:
            rn = Note(key="rounding", title="Rounding-off")
            self.p.notes.append(rn)
        line = f"All amounts are expressed in Rupees ({self.denom_word}) unless otherwise stated."
        rn.footnotes = [line] + [f for f in (rn.footnotes or []) if f != line]

    def _render_footnotes_numbered(self, ws, r, key, footnotes, start_k, policy=False):
        for j, fn in enumerate(footnotes, start=start_k):
            if policy and ":" in fn:
                head, body = fn.split(":", 1)
                self.cell(ws, r, 1, self._subnum(key, j), bold=True)
                self.cell(ws, r, 2, head.strip(), bold=True)
                r += 1
                self.cell(ws, r, 2, body.strip(), wrap=True)
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
                ws.row_dimensions[r].height = max(15 * ((len(body) // 110) + 1), 15)
                r += 1
            else:
                self.cell(ws, r, 1, self._subnum(key, j), bold=True)
                self.cell(ws, r, 2, fn, italic=True, wrap=True)
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
                ws.row_dimensions[r].height = max(15 * ((len(fn) // 110) + 1), 15)
                r += 1
        return r

    def _no_gridlines(self):
        for ws in self.wb.worksheets:
            ws.sheet_view.showGridLines = False

    def _landscape(self, ws):
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = 9
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_margins.left = ws.page_margins.right = 0.3

    # -- title block ----------------------------------------------------------
    def title_block(self, ws, desc, span=6):
        e = self.p.entity
        self.cell(ws, 1, 1, e.name, bold=True, center=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
        self.cell(ws, 2, 1, desc, bold=True, center=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
        self.cell(ws, 3, span, self._amount_caption(), italic=True, right=True)

    # -- NOTES ----------------------------------------------------------------
    def build_notes(self):
        ws = self.ws_notes
        for w, wd in {"A": 6, "B": 52, "C": 4, "D": 18, "E": 2, "F": 18}.items():
            ws.column_dimensions[w].width = wd
        self.title_block(ws, "Notes forming part of the Financial Statements")
        r = 5
        for key in self.retained:
            if key == "ppe":
                continue  # PPE keeps its own dedicated landscape sheet
            num = self.note_no[key]
            if key == "capital":
                r = self._note_capital(ws, r, num)
            elif key in ("entity", "policies", "prev_year", "rounding",
                         "segment", "related_party", "contingent", "msmed",
                         "forex", "confirmation", "dtl", "dta"):
                r = self._note_prose(ws, r, num, key)
            else:
                r = self._note_standard(ws, r, num, key)
        self.notes_last_row = r

    # -- Capital Account (rendered inline within Notes) ----------------------
    def _note_capital(self, ws, r, num):
        note = self.p.note("capital")
        title = (note.title if note and note.title else
                 ("Partners' Capital Account" if self.firm else "Owner's Capital Account"))
        self.cell(ws, r, 1, self._heading_formula("capital"), bold=True)
        self.cell(ws, r, 2, title, bold=True)
        r += 1
        if self.firm:
            r = self._cap_inline_partners(ws, r)
        else:
            r = self._cap_inline_owner(ws, r)
        if note and note.footnotes:
            r = self._render_footnotes_numbered(ws, r, "capital", note.footnotes, 1)
        return r + 1

    def _cap_inline_owner(self, ws, r):
        oc = self.p.owner_capital or OwnerCapital(name=self.p.entity.name)
        r = self._table_header(ws, r)
        add_rows, sub_rows = [], []
        for ln in oc.resolved_lines():
            self.cell(ws, r, 1, ln.prefix())
            self.cell(ws, r, 2, ln.label)
            if ln.kind == "profit":
                self.cell(ws, r, 4, "='Statement of P&L'!__NPCY__", num=True)
                self.cell(ws, r, 6, "='Statement of P&L'!__NPPY__", num=True)
            else:
                self.cell(ws, r, 4, round(ln.cy, 2), num=True)
                self.cell(ws, r, 6, round(ln.py, 2), num=True)
            (sub_rows if ln.kind in SUB_KINDS else add_rows).append(r)
            r += 1
        cl_cy = sum(l.signed("cy") for l in oc.resolved_lines())
        cl_py = sum(l.signed("py") for l in oc.resolved_lines())
        self.cell(ws, r, 2, "Closing Balance", bold=True)
        self.cell(ws, r, 4, self._closing_formula("D", add_rows, sub_rows), bold=True, num=True, top=True, val_hint=cl_cy)
        self.cell(ws, r, 6, self._closing_formula("F", add_rows, sub_rows), bold=True, num=True, top=True, val_hint=cl_py)
        self.anchor["note_capital_cy"] = f"Notes!D{r}"
        self.anchor["note_capital_py"] = f"Notes!F{r}"
        return r + 1

    def _cap_inline_partners(self, ws, r):
        cls_cy, cls_py = [], []
        for pt in self.p.partners:
            psr = f"  ({pt.psr:g}%)" if pt.psr else ""
            self.cell(ws, r, 2, f"{pt.name}{psr}", bold=True, italic=True)
            r += 1
            r = self._table_header(ws, r)
            add_rows, sub_rows = [], []
            for ln in pt.resolved_lines():
                self.cell(ws, r, 1, ln.prefix())
                self.cell(ws, r, 2, ln.label)
                self.cell(ws, r, 4, round(ln.cy, 2), num=True)
                self.cell(ws, r, 6, round(ln.py, 2), num=True)
                (sub_rows if ln.kind in SUB_KINDS else add_rows).append(r)
                r += 1
            ccy = sum(l.signed("cy") for l in pt.resolved_lines())
            cpy = sum(l.signed("py") for l in pt.resolved_lines())
            self.cell(ws, r, 2, "Closing Balance", bold=True)
            self.cell(ws, r, 4, self._closing_formula("D", add_rows, sub_rows), bold=True, num=True, top=True, val_hint=ccy)
            self.cell(ws, r, 6, self._closing_formula("F", add_rows, sub_rows), bold=True, num=True, top=True, val_hint=cpy)
            cls_cy.append(f"D{r}"); cls_py.append(f"F{r}")
            r += 2
        self.cell(ws, r, 2, "Total Partners' Capital", bold=True)
        self.cell(ws, r, 4, "=" + "+".join(cls_cy), bold=True, num=True, top=True, double_bottom=True)
        self.cell(ws, r, 6, "=" + "+".join(cls_py), bold=True, num=True, top=True, double_bottom=True)
        self.anchor["note_capital_cy"] = f"Notes!D{r}"
        self.anchor["note_capital_py"] = f"Notes!F{r}"
        return r + 1

    def _note_title(self, ws, r, num, title, with_years=True):
        self.cell(ws, r, 1, f"Note {num}", bold=True)
        self.cell(ws, r, 2, title, bold=True)
        if with_years:
            self.cell(ws, r, 4, self.p.entity.cy_label, bold=True, center=True)
            self.cell(ws, r, 6, self.p.entity.py_label, bold=True, center=True)
        return r + 1

    def _table_header(self, ws, r):
        """Shaded table-header band: Particulars + year columns (cols 1-6)."""
        for c in range(1, 7):
            self.cell(ws, r, c, None, fill=HEADER_FILL, top=True, bottom=True)
        self.cell(ws, r, 2, "Particulars", bold=True, fill=HEADER_FILL, top=True, bottom=True)
        self.cell(ws, r, 4, self.p.entity.cy_label, bold=True, center=True, fill=HEADER_FILL, top=True, bottom=True)
        self.cell(ws, r, 6, self.p.entity.py_label, bold=True, center=True, fill=HEADER_FILL, top=True, bottom=True)
        return r + 1

    def _note_standard(self, ws, r, num, key):
        note = self.p.note(key)
        if note is None:
            return r
        # heading row, kept separate from the table
        self.cell(ws, r, 1, self._heading_formula(key), bold=True)
        self.cell(ws, r, 2, note.title, bold=True)
        r += 1
        # shaded table header
        r = self._table_header(ws, r)
        first = r
        for it in note.items:
            self.cell(ws, r, 2, it.label)
            self.cell(ws, r, 4, round(it.cy, 2), num=True)
            self.cell(ws, r, 6, round(it.py, 2), num=True)
            r += 1
        last = r - 1
        self.cell(ws, r, 2, "Total", bold=True)
        if note.items:
            self.cell(ws, r, 4, f"=SUM(D{first}:D{last})", bold=True, num=True, top=True, val_hint=note.total_cy())
            self.cell(ws, r, 6, f"=SUM(F{first}:F{last})", bold=True, num=True, top=True, val_hint=note.total_py())
        else:
            self.cell(ws, r, 4, 0, bold=True, num=True, top=True)
            self.cell(ws, r, 6, 0, bold=True, num=True, top=True)
        self.anchor[f"note_{key}_cy"] = f"Notes!D{r}"
        self.anchor[f"note_{key}_py"] = f"Notes!F{r}"
        r += 1
        r = self._render_subnotes(ws, r, key, note)
        start_k = len(note.subnotes) + 1
        if note.footnotes:
            r = self._render_footnotes_numbered(ws, r, key, note.footnotes, start_k)
        return r + 1

    def _render_subnotes(self, ws, r, key, note, start_k=1):
        """Render note.subnotes as N.1, N.2 ... each with optional table/prose."""
        for k, sn in enumerate(note.subnotes, start=start_k):
            self.cell(ws, r, 1, self._subnum(key, k), bold=True)
            self.cell(ws, r, 2, sn.title, bold=True)
            r += 1
            if sn.items:
                r = self._table_header(ws, r)
                for it in sn.items:
                    self.cell(ws, r, 2, it.label)
                    self.cell(ws, r, 4, round(it.cy, 2), num=True)
                    self.cell(ws, r, 6, round(it.py, 2), num=True)
                    r += 1
            for fn in sn.footnotes:
                self.cell(ws, r, 2, fn, italic=True, wrap=True)
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
                ws.row_dimensions[r].height = max(15 * ((len(fn) // 110) + 1), 15)
                r += 1
            r += 1
        return r

    def _footnotes(self, ws, r, note):
        for fn in note.footnotes:
            c = self.cell(ws, r, 2, fn, italic=True, wrap=True)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            ws.row_dimensions[r].height = max(15 * ((len(fn) // 110) + 1), 15)
            r += 1
        return r

    def _note_prose(self, ws, r, num, key):
        note = self.p.note(key)
        title = note.title if note else self._default_title(key)
        self.cell(ws, r, 1, self._heading_formula(key), bold=True)
        self.cell(ws, r, 2, title, bold=True)
        r += 1
        if note and note.footnotes:
            body = note.footnotes
        elif note and note.items:
            body = [it.label for it in note.items]
        else:
            body = [self._default_prose(key)]
        r = self._render_footnotes_numbered(ws, r, key, body, 1, policy=(key == "policies"))
        if note and note.subnotes:
            r = self._render_subnotes(ws, r, key, note, start_k=len(body) + 1)
        return r + 1

    def _default_title(self, key):
        return {
            "entity": "Brief about the Entity",
            "policies": "Significant Accounting Policies",
            "prev_year": "Previous Year Figures",
            "rounding": "Rounding-off",
            "segment": "Segment Reporting",
            "related_party": "Related Party Disclosures",
            "contingent": "Contingent Liabilities and Commitments",
            "msmed": "MSMED Act Disclosure",
            "forex": "Earnings / Expenditure in Foreign Currency",
            "confirmation": "Confirmation of Balances",
            "dtl": "Deferred Tax Liabilities (Net)",
            "dta": "Deferred Tax Assets (Net)",
        }.get(key, key.title())

    def _default_prose(self, key):
        return {
            "rounding": "Figures have been rounded off to the nearest Rupee and presented to two decimals.",
            "segment": "The entity operates in a single business segment; accordingly, segment reporting under AS 17 is not applicable.",
            "prev_year": "Previous year figures have been regrouped / reclassified wherever necessary to conform to the current year's presentation.",
        }.get(key, "—")

    # -- Capital Account (own sheet, VERTICAL format) ------------------------
    def build_capital_sheet(self, num):
        ws = self.ws_cap
        if self.firm:
            self._landscape(ws)            # multiple partner columns -> wide
            self._capital_partners(ws, num)
        else:
            ws.page_setup.orientation = "portrait"   # single vertical account
            ws.page_setup.fitToWidth = 1
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
            self._capital_owner(ws, num)

    def _footnotes_wide(self, ws, r, note, span):
        for fn in note.footnotes:
            self.cell(ws, r, 2, fn, italic=True, wrap=True)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=span)
            ws.row_dimensions[r].height = max(15 * ((len(fn) // 130) + 1), 15)
            r += 1
        return r

    def _closing_formula(self, col, add_rows, sub_rows):
        f = "=" + "+".join(f"{col}{x}" for x in add_rows) if add_rows else "=0"
        if sub_rows:
            f += "-" + "-".join(f"{col}{x}" for x in sub_rows)
        return f

    def _capital_owner(self, ws, num):
        """Vertical proprietor capital account: every T-format particular kept
        verbatim (Add / Less prefixes), Closing computed by formula."""
        oc = self.p.owner_capital or OwnerCapital(name=self.p.entity.name)
        for w, wd in {"A": 8, "B": 48, "C": 4, "D": 18, "E": 2, "F": 18}.items():
            ws.column_dimensions[w].width = wd
        self.title_block(ws, f"Note {num} \u2014 Owner's Capital Account", span=6)
        r = 5
        for _c in range(1, 7):
            self.cell(ws, r, _c, None, fill=HEADER_FILL, top=True, bottom=True)
        self.cell(ws, r, 2, "Particulars", bold=True, fill=HEADER_FILL, top=True, bottom=True)
        self.cell(ws, r, 4, self.p.entity.cy_label, bold=True, center=True, fill=HEADER_FILL, top=True, bottom=True)
        self.cell(ws, r, 6, self.p.entity.py_label, bold=True, center=True, fill=HEADER_FILL, top=True, bottom=True)
        r += 1
        add_rows, sub_rows = [], []
        for ln in oc.resolved_lines():
            self.cell(ws, r, 1, ln.prefix())
            self.cell(ws, r, 2, ln.label)
            if ln.kind == "profit":
                self.cell(ws, r, 4, "='Statement of P&L'!__NPCY__", num=True)
                self.cell(ws, r, 6, "='Statement of P&L'!__NPPY__", num=True)
            else:
                self.cell(ws, r, 4, round(ln.cy, 2), num=True)
                self.cell(ws, r, 6, round(ln.py, 2), num=True)
            (sub_rows if ln.kind in SUB_KINDS else add_rows).append(r)
            r += 1
        cl_cy = sum(l.signed("cy") for l in oc.resolved_lines())
        cl_py = sum(l.signed("py") for l in oc.resolved_lines())
        self.cell(ws, r, 2, "Closing Balance", bold=True)
        self.cell(ws, r, 4, self._closing_formula("D", add_rows, sub_rows),
                  bold=True, num=True, top=True, val_hint=cl_cy)
        self.cell(ws, r, 6, self._closing_formula("F", add_rows, sub_rows),
                  bold=True, num=True, top=True, val_hint=cl_py)
        self.anchor["note_capital_cy"] = f"Capital Account!D{r}"
        self.anchor["note_capital_py"] = f"Capital Account!F{r}"
        r += 2
        note = self.p.note("capital") or Note(key="capital", title="Owner's Capital Account")
        if not note.footnotes:
            note.footnotes = [
                "Capital is maintained under the fluctuating capital method; a separate Current Account is not maintained.",
                "The account is presented exactly as in the books: items credited or debited "
                "directly to capital (e.g. interest income, LIC premium, personal drawings) are "
                "retained here and are not regrouped into the Statement of Profit and Loss.",
                "Closing Balance is computed by formula; Net Profit for the year is linked to the Statement of Profit and Loss."]
        self._footnotes_wide(ws, r, note, span=6)

    def _capital_partners(self, ws, num):
        """Vertical partners' capital account: one value-column-pair per partner
        plus a Total pair. Each partner keeps its own particulars verbatim."""
        parts = self.p.partners
        n = len(parts)
        # union of (label -> kind), preserving first-seen order
        kind_of, order = {}, []
        for pt in parts:
            for ln in pt.resolved_lines():
                if ln.label not in kind_of:
                    kind_of[ln.label] = ln.kind
                    order.append(ln.label)
        lut = [{ln.label: ln for ln in pt.resolved_lines()} for pt in parts]

        widths = {"A": 8, "B": 30}
        for k in range(n + 1):                      # partner pairs + total pair
            widths[get_column_letter(3 + 2 * k)] = 15
            widths[get_column_letter(4 + 2 * k)] = 15
        for w, wd in widths.items():
            ws.column_dimensions[w].width = wd
        last_col = 2 + 2 * (n + 1)
        self.title_block(ws, f"Note {num} \u2014 Partners' Capital Account", span=last_col)

        cyl, pyl = self.p.entity.cy_label, self.p.entity.py_label
        r = 5
        # super-header: partner names (+PSR) over their pairs, then Total
        col = 3
        for pt in parts:
            psr = f"  ({pt.psr:g}%)" if pt.psr else ""
            self.cell(ws, r, col, f"{pt.name}{psr}", bold=True, center=True, top=True, bottom=True)
            ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
            self.cell(ws, r, col + 1, None, top=True, bottom=True)
            col += 2
        self.cell(ws, r, col, "Total \u2014 All Partners", bold=True, center=True, top=True, bottom=True)
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
        self.cell(ws, r, col + 1, None, top=True, bottom=True)
        ws.row_dimensions[r].height = 26
        r += 1
        # sub-header: year labels under each pair
        self.cell(ws, r, 2, "Particulars", bold=True, bottom=True)
        for k in range(n + 1):
            self.cell(ws, r, 3 + 2 * k, cyl, bold=True, center=True, bottom=True)
            self.cell(ws, r, 4 + 2 * k, pyl, bold=True, center=True, bottom=True)
        r += 1
        tcol = 3 + 2 * n                            # first Total column (CY)
        add_rows, sub_rows = [], []
        for label in order:
            kind = kind_of[label]
            prefix = "" if kind == "opening" else ("Less" if kind in SUB_KINDS else "Add")
            self.cell(ws, r, 1, prefix)
            self.cell(ws, r, 2, label)
            for k, m in enumerate(lut):
                ln = m.get(label)
                self.cell(ws, r, 3 + 2 * k, round(ln.cy, 2) if ln else 0, num=True)
                self.cell(ws, r, 4 + 2 * k, round(ln.py, 2) if ln else 0, num=True)
            cy_cells = "+".join(f"{get_column_letter(3 + 2 * k)}{r}" for k in range(n))
            py_cells = "+".join(f"{get_column_letter(4 + 2 * k)}{r}" for k in range(n))
            self.cell(ws, r, tcol, f"={cy_cells}" if cy_cells else 0, num=True)
            self.cell(ws, r, tcol + 1, f"={py_cells}" if py_cells else 0, num=True)
            (sub_rows if kind in SUB_KINDS else add_rows).append(r)
            r += 1
        # closing row across every value column (partners + total)
        self.cell(ws, r, 2, "Closing Balance", bold=True)
        for ci in range(3, 3 + 2 * (n + 1)):
            L = get_column_letter(ci)
            self.cell(ws, r, ci, self._closing_formula(L, add_rows, sub_rows),
                      bold=True, num=True, top=True)
        self.anchor["note_capital_cy"] = f"Capital Account!{get_column_letter(tcol)}{r}"
        self.anchor["note_capital_py"] = f"Capital Account!{get_column_letter(tcol + 1)}{r}"
        r += 2
        note = self.p.note("capital") or Note(key="capital", title="Partners' Capital Account")
        if not note.footnotes:
            note.footnotes = [
                "Capital is maintained under the fluctuating capital method (one account per partner), presented vertically.",
                "Each partner's account shows its own particulars exactly as in the books; Closing Balances are computed by formula."]
        self._footnotes_wide(ws, r, note, span=last_col)

    # -- PPE Schedule (own landscape sheet) ----------------------------------
    def build_ppe_sheet(self, num):
        ws = self.ws_ppe
        self._landscape(ws)
        for w, wd in {"A": 4, "B": 30, "C": 14, "D": 13, "E": 14, "F": 14, "G": 13, "H": 14, "I": 14, "J": 14}.items():
            ws.column_dimensions[w].width = wd
        self.title_block(ws, f'="Note "&Notes!{self.numcell["ppe"]}&" \u2014 Property, Plant and Equipment"', span=10)
        cyl, pyl = self.p.entity.cy_label, self.p.entity.py_label
        r = 5
        # super-group header (correctly aligned to detail columns)
        self.cell(ws, r, 3, "GROSS BLOCK", bold=True, center=True, top=True, bottom=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        self.cell(ws, r, 6, "ACCUMULATED DEPRECIATION", bold=True, center=True, top=True, bottom=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
        self.cell(ws, r, 9, "NET BLOCK", bold=True, center=True, top=True, bottom=True, fill=HEADER_FILL)
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=10)
        r += 1
        heads = ["Sr.", "Particulars", f"Opening\nas at 01.04", "Additions", f"Closing\nas at {cyl}",
                 f"Opening\nas at 01.04", "For the\nyear", f"Closing\nas at {cyl}",
                 f"Net Block\n{cyl}", f"Net Block\n{pyl}"]
        for j, h in enumerate(heads):
            self.cell(ws, r, 1 + j, h, bold=True, center=(j > 1), wrap=True, top=True, bottom=True, fill=HEADER_FILL)
        ws.row_dimensions[r].height = 46
        r += 1
        first = r
        for i, a in enumerate(self.p.ppe_assets, start=1):
            self.cell(ws, r, 1, i, center=True)
            self.cell(ws, r, 2, a.name)
            self.cell(ws, r, 3, round(a.gb_open, 2), num=True)
            self.cell(ws, r, 4, round(a.additions, 2), num=True)
            self.cell(ws, r, 5, f"=C{r}+D{r}", num=True)
            self.cell(ws, r, 6, round(a.accdep_open, 2), num=True)
            self.cell(ws, r, 7, round(a.dep_year, 2), num=True)
            self.cell(ws, r, 8, f"=F{r}+G{r}", num=True)
            self.cell(ws, r, 9, f"=E{r}-H{r}", num=True)
            self.cell(ws, r, 10, f"=C{r}-F{r}", num=True)
            r += 1
        last = r - 1
        self.cell(ws, r, 2, "Total", bold=True)
        for col in (3, 4, 5, 6, 7, 8, 9, 10):
            L = get_column_letter(col)
            self.cell(ws, r, col, f"=SUM({L}{first}:{L}{last})", bold=True, num=True, top=True)
        self.anchor["note_ppe_cy"] = f"PPE Schedule!I{r}"
        self.anchor["note_ppe_py"] = f"PPE Schedule!J{r}"
        self.anchor["ppe_dep_year"] = f"PPE Schedule!G{r}"
        r += 2
        note = self.p.note("ppe") or Note(key="ppe", title="Property, Plant and Equipment")
        fns = list(note.footnotes)
        if self.p.depreciation_case == "A":
            fns.append("Depreciation has not been provided in the books; fixed-asset balances are carried at cost as gross block. The impact of non-provision has not been ascertained.")
        fns += ["No revaluation of assets was carried out during the year.",
                "There are no intangible assets, capital work-in-progress or intangible assets under development.",
                "All assets are owned by the entity and are free from charge except as disclosed."]
        note.footnotes = fns
        self._footnotes_wide(ws, r, note, span=10)

    # -- BALANCE SHEET --------------------------------------------------------
    def build_balance_sheet(self):
        ws = self.ws_bs
        for w, wd in {"A": 5, "B": 58, "C": 8, "D": 18, "E": 2, "F": 18}.items():
            ws.column_dimensions[w].width = wd
        self.title_block(ws, f"Balance Sheet as at {self.p.entity.cy_label}")
        for _c in range(1, 7):
            self.cell(ws, 4, _c, None, fill=HEADER_FILL, top=True, bottom=True)
        self.cell(ws, 4, 2, "Particulars", bold=True, fill=HEADER_FILL, top=True, bottom=True)
        self.cell(ws, 4, 3, "Note", bold=True, center=True, fill=HEADER_FILL, top=True, bottom=True)
        self.cell(ws, 4, 4, self.p.entity.cy_label, bold=True, center=True, fill=HEADER_FILL, top=True, bottom=True)
        self.cell(ws, 4, 6, self.p.entity.py_label, bold=True, center=True, fill=HEADER_FILL, top=True, bottom=True)
        r = 6
        self.cell(ws, r, 1, "I", bold=True)
        self.cell(ws, r, 2, "OWNERS' FUNDS AND LIABILITIES", bold=True)
        r += 1
        r, liab_rows = self._bs_block(ws, r, BS_LIABILITIES)
        self.cell(ws, r, 2, "Total", bold=True)
        self.cell(ws, r, 4, "=" + ("+".join(f"D{x}" for x in liab_rows) or "0"), bold=True, num=True, top=True, double_bottom=True)
        self.cell(ws, r, 6, "=" + ("+".join(f"F{x}" for x in liab_rows) or "0"), bold=True, num=True, top=True, double_bottom=True)
        self.anchor["bs_liab_total_cy"] = f"Balance Sheet!D{r}"
        r += 2
        self.cell(ws, r, 1, "II", bold=True)
        self.cell(ws, r, 2, "ASSETS", bold=True)
        r += 1
        r, asset_rows = self._bs_block(ws, r, BS_ASSETS)
        self.cell(ws, r, 2, "Total", bold=True)
        self.cell(ws, r, 4, "=" + ("+".join(f"D{x}" for x in asset_rows) or "0"), bold=True, num=True, top=True, double_bottom=True)
        self.cell(ws, r, 6, "=" + ("+".join(f"F{x}" for x in asset_rows) or "0"), bold=True, num=True, top=True, double_bottom=True)
        r += 2
        self.cell(ws, r, 2, "The accompanying notes are an integral part of the financial statements.", italic=True)

    def _bs_block(self, ws, r, plan):
        subtotal_rows = []
        for header, lines in plan:
            visible = [(lbl, key) for (lbl, key) in lines if key in self.note_no or key == "ppe"]
            visible = [(lbl, key) for (lbl, key) in lines
                       if (key in self.note_no) or (key == "ppe" and "note_ppe_cy" in self.anchor)]
            if not visible:
                continue
            self.cell(ws, r, 1, header.split()[0], bold=True)
            self.cell(ws, r, 2, header[len(header.split()[0]):].strip(), bold=True)
            r += 1
            block_rows = []
            for lbl, key in lines:
                cy_anchor = self.anchor.get(f"note_{key}_cy")
                if cy_anchor is None:
                    continue
                self.cell(ws, r, 2, lbl)
                if key in self.note_no:
                    self.cell(ws, r, 3, f"=Notes!{self.numcell[key]}", center=True)
                self.cell(ws, r, 4, f"='{self._sheet(cy_anchor)}'!{self._addr(cy_anchor)}", num=True)
                py_anchor = self.anchor.get(f"note_{key}_py")
                self.cell(ws, r, 6, f"='{self._sheet(py_anchor)}'!{self._addr(py_anchor)}", num=True)
                block_rows.append(r)
                r += 1
            # subtotal
            if block_rows:
                self.cell(ws, r, 2, "", )
                self.cell(ws, r, 4, "=" + "+".join(f"D{x}" for x in block_rows), num=True, top=True, bold=True)
                self.cell(ws, r, 6, "=" + "+".join(f"F{x}" for x in block_rows), num=True, top=True, bold=True)
                subtotal_rows.append(r)
                r += 1
        return r, subtotal_rows

    @staticmethod
    def _sheet(anchor):
        return anchor.split("!")[0]

    @staticmethod
    def _addr(anchor):
        return anchor.split("!")[1]

    # -- STATEMENT OF P&L -----------------------------------------------------
    def _note_sum(self, key, yr):
        n = self.p.note(key)
        return round(sum((getattr(i, yr) or 0) for i in n.items), 2) if n else 0.0

    def _pl_profit(self, yr):
        inc = self._note_sum("revenue", yr) + self._note_sum("other_income", yr)
        exp = sum(self._note_sum(k, yr) for k in ("cost_materials", "changes_inventory",
                  "employee_benefits", "finance_costs", "depreciation", "other_expenses"))
        return round(inc - exp, 2)

    def build_pl(self):
        ws = self.ws_pl
        for w, wd in {"A": 5, "B": 58, "C": 8, "D": 18, "E": 2, "F": 18}.items():
            ws.column_dimensions[w].width = wd
        self.title_block(ws, f"Statement of Profit and Loss for the year ended {self.p.entity.cy_label}")
        for _c in range(1, 7):
            self.cell(ws, 4, _c, None, fill=HEADER_FILL, top=True, bottom=True)
        self.cell(ws, 4, 2, "Particulars", bold=True, fill=HEADER_FILL, top=True, bottom=True)
        self.cell(ws, 4, 3, "Note", bold=True, center=True, fill=HEADER_FILL, top=True, bottom=True)
        self.cell(ws, 4, 4, f"Year ended {self.p.entity.cy_label}", bold=True, center=True, fill=HEADER_FILL, top=True, bottom=True)
        self.cell(ws, 4, 6, f"Year ended {self.p.entity.py_label}", bold=True, center=True, fill=HEADER_FILL, top=True, bottom=True)
        r = 6

        def line(roman, label, key=None, val_cy=None, val_py=None, bold=False, top=False, hint_cy=None, hint_py=None):
            nonlocal r
            self.cell(ws, r, 1, roman, bold=bold)
            self.cell(ws, r, 2, label, bold=bold)
            if key and key in self.note_no:
                self.cell(ws, r, 3, f"=Notes!{self.numcell[key]}", center=True)
            if val_cy is not None:
                self.cell(ws, r, 4, val_cy, num=True, bold=bold, top=top, val_hint=hint_cy)
            if val_py is not None:
                self.cell(ws, r, 6, val_py, num=True, bold=bold, top=top, val_hint=hint_py)
            rr = r
            r += 1
            return rr

        rev_cy = self._anchor_ref("revenue", "cy")
        rev_py = self._anchor_ref("revenue", "py")
        oi_cy = self._anchor_ref("other_income", "cy")
        oi_py = self._anchor_ref("other_income", "py")
        r_rev = line("I", "Revenue from operations", "revenue", rev_cy, rev_py)
        r_oi = line("II", "Other income", "other_income", oi_cy, oi_py)
        r_ti = line("III", "Total Income (I + II)", None,
                    f"=D{r_rev}+D{r_oi}", f"=F{r_rev}+F{r_oi}", bold=True, top=True)
        line("IV", "Expenses:", None, bold=True)
        exp_rows = []
        for lbl, key in PL_EXPENSE_PLAN:
            if key not in self.note_no:
                continue
            cy = self._anchor_ref(key, "cy")
            py = self._anchor_ref(key, "py")
            rr = line("", lbl, key, cy, py)
            exp_rows.append(rr)
        r_te = line("", "Total expenses", None,
                    "=" + "+".join(f"D{x}" for x in exp_rows) if exp_rows else "0",
                    "=" + "+".join(f"F{x}" for x in exp_rows) if exp_rows else "0",
                    bold=True, top=True)
        prof_cy, prof_py = self._pl_profit("cy"), self._pl_profit("py")
        r_pbt = line("V", "Profit before tax (III − IV)", None,
                     f"=D{r_ti}-D{r_te}", f"=F{r_ti}-F{r_te}", bold=True, top=True,
                     hint_cy=prof_cy, hint_py=prof_py)
        if self.firm:
            tax = self.p.firm_tax
            r_tax = line("VI", "Tax expense — Current tax", "st_provisions",
                         round(tax.current_tax_cy, 2) if tax else 0,
                         round(tax.current_tax_py, 2) if tax else 0)
            tcy = round(tax.current_tax_cy, 2) if tax else 0
            tpy = round(tax.current_tax_py, 2) if tax else 0
            r_pat = line("VII", "Profit for the year after tax (V − VI)", None,
                         f"=D{r_pbt}-D{r_tax}", f"=F{r_pbt}-F{r_tax}", bold=True, top=True,
                         hint_cy=round(prof_cy - tcy, 2), hint_py=round(prof_py - tpy, 2))
            self.anchor["np_cy"] = f"Statement of P&L!D{r_pat}"
            self.anchor["np_py"] = f"Statement of P&L!F{r_pat}"
            self._appropriation(ws, r, r_pat)
        else:
            self.cell(ws, r_pbt, 2, "Profit for the year (= V)", bold=True)
            self.anchor["np_cy"] = f"Statement of P&L!D{r_pbt}"
            self.anchor["np_py"] = f"Statement of P&L!F{r_pbt}"
        self._wire_capital_profit()

    def _appropriation(self, ws, r, r_pat):
        r += 1
        self.cell(ws, r, 2, "Profit and Loss Appropriation", bold=True); r += 1
        self.cell(ws, r, 2, "Profit for the year after tax")
        self.cell(ws, r, 4, f"=D{r_pat}", num=True)
        self.cell(ws, r, 6, f"=F{r_pat}", num=True)
        r += 1
        int_cy = sum(p.kind_total("interest", "cy") for p in self.p.partners)
        int_py = sum(p.kind_total("interest", "py") for p in self.p.partners)
        rem_cy = sum(p.kind_total("remuneration", "cy") for p in self.p.partners)
        rem_py = sum(p.kind_total("remuneration", "py") for p in self.p.partners)
        self.cell(ws, r, 2, "(−) Interest on partners' capital")
        self.cell(ws, r, 4, round(int_cy, 2), num=True)
        self.cell(ws, r, 6, round(int_py, 2), num=True); r_int = r; r += 1
        self.cell(ws, r, 2, "(−) Remuneration to partners")
        self.cell(ws, r, 4, round(rem_cy, 2), num=True)
        self.cell(ws, r, 6, round(rem_py, 2), num=True); r_rem = r; r += 1
        self.cell(ws, r, 2, "Divisible profit among partners", bold=True)
        self.cell(ws, r, 4, f"=D{r_pat}-D{r_int}-D{r_rem}", bold=True, num=True, top=True)
        self.cell(ws, r, 6, f"=F{r_pat}-F{r_int}-F{r_rem}", bold=True, num=True, top=True)
        r += 1

    def _anchor_ref(self, key, yr):
        a = self.anchor.get(f"note_{key}_{yr}")
        if not a:
            return 0
        return f"='{self._sheet(a)}'!{self._addr(a)}"

    def _wire_capital_profit(self):
        """Replace capital-account profit placeholders with real P&L links."""
        # Owner capital placeholders were written as formulas with __NPCY__/__NPPY__
        np_cy = self.anchor.get("np_cy")
        np_py = self.anchor.get("np_py")
        if not self.firm and np_cy and getattr(self, "ws_notes", None) is not None:
            ws = self.ws_notes
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and "__NPCY__" in c.value:
                        c.value = f"='{self._sheet(np_cy)}'!{self._addr(np_cy)}"
                    elif isinstance(c.value, str) and "__NPPY__" in c.value:
                        c.value = f"='{self._sheet(np_py)}'!{self._addr(np_py)}"

    # -- INDEX ----------------------------------------------------------------
    def build_index(self):
        ws = self.ws_idx
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 12
        self.title_block(ws, "Index")
        self.cell(ws, 4, 2, "Contents", bold=True)
        self.cell(ws, 4, 3, "Reference", bold=True, center=True)
        r = 5
        nav = [("Balance Sheet", "Balance Sheet"),
               ("Statement of Profit and Loss", "Statement of P&L"),
               ("Notes to the Financial Statements", "Notes")]
        if getattr(self, "ws_ppe", None) is not None:
            nav.append((f"Note {self.note_no['ppe']} - Property, Plant and Equipment", "PPE Schedule"))
        for label, sheet in nav:
            c = self.cell(ws, r, 2, label)
            c.hyperlink = f"#'{sheet}'!A1"
            c.font = self._f(color="0563C1")
            self.cell(ws, r, 3, "→", center=True)
            r += 1
        r += 1
        self.cell(ws, r, 2, "Note index", bold=True); r += 1
        for key in self.retained:
            _lbl = self._note_label(key).replace(chr(34), "'")
            self.cell(ws, r, 2, f'="Note "&Notes!{self.numcell[key]}&"  {_lbl}"')
            r += 1

    def _note_label(self, key):
        note = self.p.note(key)
        if note and note.title:
            return note.title
        if key == "capital":
            return "Partners' Capital Account" if self.firm else "Owner's Capital Account"
        if key == "ppe":
            return "Property, Plant and Equipment"
        return self._default_title(key)

    # -- save with theme/styles font patch -----------------------------------
    def _save_with_theme_patch(self) -> bytes:
        buf = io.BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return self._patch_fonts(buf.read())

    @staticmethod
    def _patch_fonts(data: bytes) -> bytes:
        """Force Calibri Light onto the theme minor font and the default style
        so empty cells comply too."""
        src = io.BytesIO(data)
        out = io.BytesIO()
        with zipfile.ZipFile(src, "r") as zin:
            names = zin.namelist()
            with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
                for n in names:
                    content = zin.read(n)
                    if n == "xl/theme/theme1.xml":
                        t = content.decode("utf-8")
                        t = re.sub(r'(<a:minorFont>\s*<a:latin typeface=")[^"]*"',
                                   r'\g<1>Calibri Light"', t)
                        t = re.sub(r'(<a:majorFont>\s*<a:latin typeface=")[^"]*"',
                                   r'\g<1>Calibri Light"', t)
                        content = t.encode("utf-8")
                    zout.writestr(n, content)
        return out.getvalue()


def build_workbook(payload: Payload, denomination: str = "actual") -> bytes:
    return Engine(payload, denomination).build()
